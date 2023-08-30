
# Đếm số file pdf trong 1 thư mục
# find . f -name "*.pdf" | wc -l
# find . -iname "*.pdf" > quynhlam.sh

import os
import openpyxl
from PyPDF4 import PdfFileReader
from datetime import datetime

 

# Hàm kiểm tra trang pdf có phải là A3 không và trả về vị trí của trang A3 nếu có
def check_a3_page(pdf_reader, page_num):
    page = pdf_reader.getPage(page_num)
    page_edges = page.cropBox
    left_edge = page_edges.getLowerLeft()
    right_edge = page_edges.getLowerRight()
    top_edge = page_edges.getUpperLeft()
    bottom_edge = page_edges.getLowerLeft()
    
    width = abs(right_edge[0] - left_edge[0])
    height = abs(top_edge[1] - bottom_edge[1])
    
    #print  (f"trang: {page_num+1}, size: {width} x {height}, sub: {width-height} ")
    #if (width >= 600 and height >= 900 and width > height) or (width >= 900 and height >= 600 and width < height):
    if abs(width - height) >= 300:
        return page_num+1
    else:
        return None

# Hàm đếm số trang A3 của file pdf và trả về một list vị trí của các trang A3
def count_a3_pages(file_path):
    with open(file_path, 'rb') as f:
        pdf_reader = PdfFileReader(f)
        page_count = pdf_reader.getNumPages()
        a3_count = 0
        a3_pages = []
        for page_num in range(page_count):
            a3_page_num = check_a3_page(pdf_reader, page_num)
            if a3_page_num is not None:
                a3_count += 1
                a3_pages.append(a3_page_num)
        return a3_count, a3_pages


# Đường dẫn tuyệt đối đến thư mục chứa tất cả các tệp PDF cần xử lý

#pdf_folder = '/mnt/g/Shared drives/SCAN/CongThuong-HoaBinh'

pdf_folder = '/mnt/d/03. Phường Đồng Tiến'
#pdf_folder = '/mnt/d/04. Phường Thịnh Lang'

print(f'Bắt đầu đếm A3 thư mục: {pdf_folder}')
print('Điều kiện thoả trang A3 là: abs(width - height) >= 300')
print('')

# Tạo workbook mới
workbook = openpyxl.Workbook()
worksheet = workbook.active
worksheet.title = 'Page Count'

# Ghi header cho bảng
worksheet.cell(row=1, column=1, value='File Name')
worksheet.cell(row=1, column=2, value='Total Pages')
worksheet.cell(row=1, column=3, value='A3 Page Count')
worksheet.cell(row=1, column=4, value='A3 Page Position')

# total_a3_pages = 0
# total_total_pages = 0

def process_pdf_files(pdf_folder):
    row_num = 2  # Bắt đầu ghi từ hàng thứ 2
    total_a3_pages = 0  # Khởi tạo tổng số trang A3
    total_pages = 0
    total_total_pages = 0
    # đối tượng datetime chứa ngày và giờ hiện tại    
    hientai = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    for root, dirs, files in os.walk(pdf_folder):
        for file in files:
            if file.endswith('.pdf'):
                file_path = os.path.join(root, file)
                try:
                    with open(file_path, 'rb') as f:
                        pdf_reader = PdfFileReader(f)
                        total_pages = pdf_reader.getNumPages()
                        a3_count, a3_pages = count_a3_pages(file_path)

                        total_a3_pages += a3_count
                        total_total_pages += total_pages

                        # Ghi dữ liệu vào bảng
                        worksheet.cell(row=row_num, column=1, value=os.path.basename(file_path))
                        worksheet.cell(row=row_num, column=2, value=total_pages)
                        worksheet.cell(row=row_num, column=3, value=a3_count)
                        worksheet.cell(row=row_num, column=4, value=','.join(map(str, a3_pages)))

                        # In thông tin số trang A3 của từng tập tin
                        print(f'{os.path.basename(file_path)}, A3: {a3_count}/{total_pages} , Vị trí A3: {a3_pages}')

                        row_num += 1
                except Exception as e:
                    print(f"Error processing file {file_path}: {e}")
                    continue
    # Ghi dữ liệu Tong vào cuoi bảng
    worksheet.cell(row=row_num, column=1, value="Tổng")
    worksheet.cell(row=row_num, column=2, value=total_total_pages)
    worksheet.cell(row=row_num, column=3, value=total_a3_pages)
    worksheet.cell(row=row_num, column=4, value=hientai)    
    # Tong quy ra A4
    worksheet.cell(row=row_num+1, column=2, value=(total_total_pages+total_a3_pages))

    # In thông tin tổng số trang A3 của toàn bộ thư mục
    print(f'Total A3 pages/ Total pages: {total_a3_pages}/{total_total_pages} = {total_a3_pages+total_total_pages}')
    

process_pdf_files(pdf_folder)

workbook.save('exel/DongTien2022.xlsx')
print("File được lưu tại: exel/DongTien2022.xlsx")

# workbook.save('exel/ThinhLang2022.xlsx')
# print("File được lưu tại: exel/ThinhLang2022.xlsx")